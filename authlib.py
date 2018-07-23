import openpyxl
import random


class AuthHandler:
    def __init__(self, db_path):
        self._auth_queue = {}
        self._auth_num = 0
        self._db_path = db_path
        self._init_db()

    def _init_db(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        labels = ['id', 'username', 'name', 'job', 'interests', 'is_active']
        ws.append(labels)

        # test labels
        labels = [1, 'username1', 'name1', 'job1', 'interests1', True]
        ws.append(labels)
        self._auth_num += 1
        labels = [2, 'username2', 'name2', 'job2', 'interests2', True]
        ws.append(labels)
        self._auth_num += 1

        wb.save(filename=self._db_path)

    def _increment_step(self, client_id):
        self._auth_queue[client_id]['step'] += 1

    def _remove_client(self, client_id):
        self._auth_queue.pop(client_id)

    def _get_step(self, client_id):
        return self._auth_queue[client_id]['step']

    def _append_data(self, client_id, data):
        self._auth_queue[client_id]['data'].append(data)

    def _add_to_db(self, client_id):
        wb = openpyxl.load_workbook(filename=self._db_path)
        ws = wb['Participants']

        r = 0
        for i in range(2, self._auth_num + 2):
            cell = 'A' + str(i)
            if ws[cell].value == client_id:
                r = i
                break

        if r == 0:
            ws.append(self._auth_queue[client_id]['data'])
        else:
            index = 0
            for c in 'ABCDEF':
                cell = c + str(r)
                ws[cell].value = self._auth_queue[client_id]['data'][index]
                index += 1

        self._auth_num += 1
        wb.save(self._db_path)

    def get_auth_num(self):
        return self._auth_num

    def add_client(self, client_id):
        self._auth_queue[client_id] = {
            'step': 0,
            'data': []
        }

    def is_in_queue(self, client_id):
        """
        Возвращает True, если пользователь еще не делал запрос на авторизацию,
        либо уже авторизировался, False - иначе
        """
        return self._auth_queue.get(client_id) is None

    def make_step(self, client_id, message, bot):
        """
        Прим.: на шаге 0 обрабатывается сообщение от бота, на последующих - от пользователя
        """
        step = self._get_step(client_id)
        if step == 0:
            self._append_data(client_id, message.chat.id)
            self._increment_step(client_id)
            bot.send_message(message.chat.id, "Как тебя зовут?")
        elif step == 1:
            self._append_data(client_id, message.from_user.username)
            self._append_data(client_id, message.text)
            self._increment_step(client_id)
            bot.send_message(message.chat.id, "Кем ты работаешь?")
        elif step == 2:
            self._append_data(client_id, message.text)
            self._increment_step(client_id)
            bot.send_message(message.chat.id, "Расскажи немного о себе и своих интересах")
        elif step == 3:
            self._append_data(client_id, message.text)
            self._append_data(client_id, True)

            self._add_to_db(client_id)
            self._remove_client(client_id)
            bot.send_message(message.chat.id, "Спасибо! Я записал тебя в список участников")

    def get_profile(self, client_id):
        wb = openpyxl.load_workbook(filename=self._db_path)
        ws = wb['Participants']

        r = 0
        for i in range(2, self._auth_num + 2):
            cell = 'A' + str(i)
            if ws[cell].value == client_id:
                r = i
                break

        if r == 0:
            return None
        else:
            result = []
            for c in 'CDE':
                cell = c + str(r)
                result.append(ws[cell].value)
            return result

    def is_authorized(self, client_id):
        wb = openpyxl.load_workbook(filename=self._db_path)
        ws = wb['Participants']

        for i in range(2, self._auth_num + 2):
            cell = 'A' + str(i)
            if ws[cell].value == client_id:
                return True
        return False

    def get_participant(self, client_id):
        wb = openpyxl.load_workbook(filename=self._db_path)
        ws = wb['Participants']

        possible = []  # номера рядов потенциальных собеседников
        for i in range(2, self._auth_num + 2):
            if ws['A' + str(i)].value != client_id and ws['F' + str(i)].value:
                possible.append(i)

        try:
            r = random.choice(possible)
            result = []
            for c in 'ABCDE':
                cell = c + str(r)
                result.append(ws[cell].value)
            return result
        except IndexError:
            return None


    def get_users(self):  # for test only
        wb = openpyxl.load_workbook(filename=self._db_path)
        ws = wb['Participants']

        result = []
        for row in ws.rows:
            result.append([cell.value for cell in row])

        return result[1:]
